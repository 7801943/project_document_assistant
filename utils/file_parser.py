import os
from pathlib import Path
from typing import List, Optional, Tuple, Union # 添加类型提示
import pdfplumber
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from docx import Document  # ✅ 用于加载文档
from docx.document import Document as DocxDocument  # ✅ 仅用于类型判断
from docx.table import _Cell
from docx.text.paragraph import Paragraph
from docx.table import Table


from loguru import logger # 使用 Loguru logger

# 注意：Loguru logger 通常在主应用程序 (例如 server_v6.py) 中配置。
# file_parser.py 直接使用导入的 logger 实例。

def get_xlsx_sheet_names(file_path: Union[str,Path]) -> List[str]:
    """
    获取 XLSX 文件中所有工作表的名称。
    参数:
        file_path (str): XLSX 文件的路径。
    返回:
        List[str]: 工作表名称的列表。如果发生错误则返回空列表。
    """

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return workbook.sheetnames
    except InvalidFileException:
        logger.error(f"文件 '{file_path}' 不是有效的XLSX文件或已损坏 (在获取sheet名时)。")
        return []
    except Exception as e:
        logger.error(f"获取 XLSX 文件 '{file_path}' 的 sheet 名称失败: {e}", exc_info=True)
        return []

def parse_xlsx_sheet_content(
    file_path: Union[str,Path],
    sheet_name: str,
    column_config: Optional[Tuple[Optional[int], Optional[int]]] = None, # 允许元组中的元素为None
    cell_delimiter: str = "\t"
) -> List[str]:
    """
    解析XLSX文件的特定工作表内容。

    参数:
        file_path (str): XLSX文件的路径。
        sheet_name (str): 要解析的工作表的名称。
        column_config (Optional[Tuple[Optional[int], Optional[int]]]):
            一个元组 (min_col, max_col)，指定要读取的列范围 (1-based)。
            如果元组中的某个值为 None，则表示该边界不限制。
            如果整个 column_config 为 None，则读取所有列。
        cell_delimiter (str): 用于连接一行中各单元格内容的分隔符。

    返回:
        List[str]: 一个字符串列表，每个字符串代表工作表的一行。
                   如果发生错误或sheet未找到，则返回空列表。
    """
    lines: List[str] = []
    try:
        # 使用 data_only=True 来获取单元格的计算值而不是公式
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' 在文件 '{file_path}' 中未找到。可用 Sheets: {', '.join(workbook.sheetnames)}")
            return []
        sheet = workbook[sheet_name]

        # iter_rows 的 min_col, max_col 参数是 1-based
        current_min_col, current_max_col = None, None
        if column_config:
            current_min_col, current_max_col = column_config

        # sheet.iter_rows() 可以接受 min_col 和 max_col 参数
        for row_cells_obj in sheet.iter_rows(min_col=current_min_col, max_col=current_max_col):
            row_values = [str(cell.value) if cell.value is not None else "" for cell in row_cells_obj]
            lines.append(cell_delimiter.join(row_values))
        return lines
    except InvalidFileException:
        logger.error(f"文件 '{file_path}' 不是有效的XLSX文件或已损坏 (在解析sheet内容时)。")
        return []
    except Exception as e:
        logger.error(f"解析 XLSX 文件 '{file_path}' (Sheet: '{sheet_name}') 时发生错误: {e}", exc_info=True)
        return []


def parse_pdf(file_path: Union[str,Path], table_delimiter: str = "\t", max_pages: int = 500) -> Optional[str]:
    """
    使用 pdfplumber 解析 PDF 文件，提取文本和表格信息。
    此版本尽量保持原始逻辑，错误输出改用 Loguru。

    参数:
        file_path (str): PDF 文件的路径。
        table_delimiter (str): 表格行内单元格之间的分隔符。
        max_pages (int): 最大处理页数。

    返回:
        Optional[str]: 解析后的文本内容，或在失败时返回 None。
    """
    text_content: List[str] = []
    # pdfplumber 依赖 pdfminer.six，后者可能产生一些 INFO 级别的日志。
    # 如果 server_v6.py 中的 Loguru 配置级别高于 INFO，这些日志可能不会显示。
    # 若要完全静默 pdfminer，可以取消下一行的注释，但这通常由主应用的日志配置控制。
    # logging.getLogger("pdfminer").setLevel(logging.ERROR) # 已在原始代码中，保持
    try:
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= max_pages:
                    logger.info(f"PDF '{file_path}'：达到最大页数限制 {max_pages}，停止解析。")
                    break
                # 提取文本
                try:
                    page_text = page.extract_text() or ""
                    text_content.append(page_text)
                except Exception as e_text:
                    logger.warning(f"解析 PDF 页面文本失败: '{file_path}', 页码: {i+1}, 错误: {e_text}")
                    continue # 继续处理下一页或下一个表格

                # 提取表格
                try:
                    tables = page.extract_tables()
                    if tables: # 确保 tables 不是 None 或空列表
                        for table in tables:
                            if table: # 确保 table 不是 None
                                for row in table:
                                    if row:  # 跳过空行 (row 本身是列表，列表为空也会跳过)
                                        text_content.append(table_delimiter.join(str(cell) if cell is not None else "" for cell in row))
                except Exception as e_table:
                    logger.warning(f"解析 PDF 表格失败: '{file_path}', 页码: {i+1}, 错误: {e_table}")
                    continue # 继续处理下一页

        content = "\n".join(text_content)
        return content
    except Exception as e_main:
        logger.error(f"解析 PDF 文件 '{file_path}' 失败: {e_main}", exc_info=True)
        return None


def parse_docx(file_path: Union[str,Path], table_delimiter: str = "\t") -> Optional[str]:
    """
    使用 python-docx 解析 docx 文件，保证段落和表格顺序与原文档一致。
    此版本尽量保持原始逻辑，错误输出改用 Loguru。

    参数:
        file_path (str): DOCX 文件的路径。
        table_delimiter (str): 表格行内单元格之间的分隔符。

    返回:
        Optional[str]: 解析后的文本内容，或在失败时返回 None。
    """
    def iter_block_items(parent: Union[DocxDocument, _Cell]):
        """
        迭代文档或单元格中的所有顶级块级元素 (段落和表格)。
        """
        if isinstance(parent, DocxDocument):
            parent_elm = parent.element.body
        elif isinstance(parent, _Cell): # 处理嵌套表格的情况
            parent_elm = parent._tc
        else:
            logger.error(f"iter_block_items 收到不支持的父元素类型: {type(parent)}")
            return # 或者 raise TypeError

        for child in parent_elm.iterchildren():
            if child.tag.endswith('p'): # 段落
                yield Paragraph(child, parent)
            elif child.tag.endswith('tbl'): # 表格
                yield Table(child, parent)

    text_content: List[str] = []
    try:
        doc = Document(file_path)
        for block in iter_block_items(doc):
            if isinstance(block, Paragraph):
                text_content.append(block.text) # 原始逻辑：直接获取段落文本
            elif isinstance(block, Table):
                for row in block.rows:
                    # 原始逻辑：替换单元格内的换行符为制表符
                    row_cells = [cell.text.replace("\n", "\t") for cell in row.cells]
                    text_content.append(table_delimiter.join(row_cells))
        content = "\n".join(text_content)
        return content
    except Exception as e:
        logger.error(f"解析 DOCX 文件 '{file_path}' 失败: {e}", exc_info=True)
        return None


def parse_xlsx(file_path: Union[str,Path], cell_delimiter: str = "\t") -> Optional[str]:
    """
    使用 openpyxl 解析 xlsx 文件，提取所有 sheet 的文本。
    此函数为 `parse_file` 的通用 XLSX 解析器，保持原始的列限制逻辑。
    错误输出改用 Loguru。

    参数:
        file_path (str): XLSX 文件的路径。
        cell_delimiter (str): 单元格内容连接的分隔符。

    返回:
        Optional[str]: 解析后的文本内容，或在失败时返回 None。
    """
    text_content: List[str] = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True) # read_only=True 提高性能
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_content.append(f"=== Sheet: {sheet_name} ===")
            # 保持原始代码中的列限制 (min_col=1, max_col=6)
            for row in sheet.iter_rows(values_only=True, min_col=1, max_col=6):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                text_content.append(cell_delimiter.join(row_values))
        content = "\n".join(text_content)
        return content
    except InvalidFileException:
        logger.error(f"文件 '{file_path}' 不是有效的XLSX文件或已损坏 (在 parse_xlsx 中)。")
        return None
    except Exception as e:
        logger.error(f"解析 XLSX 文件 '{file_path}' 失败: {e}", exc_info=True)
        return None


def parse_file(file_path: Union[str,Path], delimiter: str = "\t") -> Optional[str]:
    """
    根据文件扩展名，自动调用相应的解析函数。

    参数:
        file_path (str): 待解析文件的路径。
        delimiter (str): 用于表格或行内数据分隔的字符。

    返回:
        Optional[str]: 解析后的文本内容。
                       如果文件不存在、不支持或解析失败，则返回包含错误信息的字符串或 None。
    """
    ext = os.path.splitext(file_path)[1].lower()
    # logger.debug(f"开始解析文件: '{file_path}' (类型: {ext}), 分隔符: '{delimiter}'")
    content: Optional[str] = None

    if not os.path.exists(file_path):
        logger.error(f"文件未找到: {file_path}")
        return f"错误: 文件 {os.path.basename(file_path)} 未找到。"
    if not os.path.isfile(file_path): # 确保是文件而不是目录
        logger.error(f"路径不是一个文件: {file_path}")
        return f"错误: 路径 {os.path.basename(file_path)} 不是一个有效文件。"

    try:
        if ext == ".pdf":
            content = parse_pdf(file_path, table_delimiter=delimiter)
        elif ext == ".xlsx":
            # parse_xlsx 用于 get_file_content 的通用解析
            # 特定sheet的读取由 compare_project_file 等工具直接调用 parse_xlsx_sheet_content
            content = parse_xlsx(file_path, cell_delimiter=delimiter)
        elif ext == ".docx":
            content = parse_docx(file_path, table_delimiter=delimiter)
        # 增加对常见纯文本格式的直接读取
        elif ext in ['.txt', '.md', '.csv', '.log', '.json', '.xml', '.html', '.yaml', '.yml', '.ini', '.cfg', '.py', '.js', '.ts', '.java', '.c', '.cpp', '.h', '.hpp', '.cs', '.go', '.php', '.rb', '.sh', '.bat']:
            logger.info(f"文件 '{file_path}' (类型: {ext}) 将作为纯文本文件读取。")
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            except Exception as e_txt:
                logger.error(f"作为纯文本文件读取 '{file_path}' 失败: {e_txt}")
                content = f"错误: 文件 {os.path.basename(file_path)} 作为纯文本读取失败。" # 返回错误信息
        else:
            logger.warning(f"文件 '{file_path}' 的扩展名 '{ext}' 不在已知解析类型中，也非标准纯文本类型。")
            content = f"错误: 不支持的文件类型 {ext}。" # 返回错误信息

    except Exception as e_parse: # 捕获在调用解析函数时可能发生的其他意外错误
        logger.error(f"解析文件 '{file_path}' 过程中发生顶层错误: {e_parse}", exc_info=True)
        content = f"错误: 解析文件 {os.path.basename(file_path)} 时发生严重错误。"

    # 在返回前最后检查 content
    if content is None:
        # 如果特定解析器返回 None (通常表示其内部已记录错误)
        logger.warning(f"文件 '{file_path}' 的特定解析器返回 None。")
        return f"错误: 文件 {os.path.basename(file_path)} 解析失败或不受支持。"

    # 如果 content 已经是错误消息字符串，直接返回
    if isinstance(content, str) and content.startswith("错误:"):
        # logger.debug(f"文件 '{file_path}' 解析返回预定义的错误消息: {content}") # 此日志可能冗余
        return content

    # logger.debug(f"文件 '{file_path}' 解析完成，内容长度: {len(content)}")
    return content
